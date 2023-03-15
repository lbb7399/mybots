import pickle
from parallelHillClimber import PARALLEL_HILL_CLIMBER
from numpy.random import SeedSequence, default_rng
import constants as c
import numpy as np
from statistics import median,mean
from scipy.stats import ttest_ind
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
import matplotlib.pyplot as plt






for i, runNum in enumerate(c.runnumberstot):
    print(runNum)
    maxFitness = np.zeros(c.numberOfGenerations+1)
    minFitness = np.zeros(c.numberOfGenerations+1)
    medFitness = np.zeros(c.numberOfGenerations+1)
    avgFitness = np.zeros(c.numberOfGenerations+1)
    allFitness = {}
    for i in range(c.populationSize):
        allFitness[i] = []

    for currentGeneration in range(c.numberOfGenerations+1):
        currentGen = currentGeneration
        parents = pickle.load( open( f"pickles/{runNum}/save_parents{runNum}-{currentGen}.p", "rb" ) )
        fitnesses = []
        for i, key in enumerate(parents.keys()):
            if currentGen != 0:
                allFitness[key].append(parents[key].fitness)
            fitnesses.append(parents[key].fitness)
            if i == 0:
                bestfit = parents[key].fitness
                worstfit = parents[key].fitness
                bestfitkey = key
                worstfitkey = key
            if parents[key].fitness < bestfit:
                bestfit = parents[key].fitness
                bestfitkey = key
            if parents[key].fitness > worstfit:
                worstfit = parents[key].fitness
                worstfitkey = key

        medfit = median(fitnesses)
        avgfit = mean(fitnesses)

        maxFitness[currentGen] = -1*bestfit
        minFitness[currentGen] = -1*worstfit
        medFitness[currentGen] = -1*medfit
        avgFitness[currentGen] = -1*avgfit
    maxfilename = f"data/maxfitnesscurve{runNum}.npy"
    minfilename = f"data/minfitnesscurve{runNum}.npy"
    medianfilename = f"data/medianfitnesscurve{runNum}.npy"
    avgfilename = f"data/avgfitnesscurve{runNum}.npy"
    np.save(maxfilename, maxFitness)
    np.save(minfilename, minFitness)
    np.save(medianfilename, medFitness)
    np.save(avgfilename, avgFitness)
    
    muts = pickle.load( open( f"pickles/{runNum}/save_mutations{runNum}-{c.numberOfGenerations}.p", "rb" ) )
    for i in range(c.populationSize):
        muts[f"{i}Fit"] = allFitness[i]
    
    pickle.dump(muts, open(f"pickles/{runNum}/save_mutations{runNum}-{c.numberOfGenerations}-fit.p", "wb"))
    




rand_fitness = []
rand_best_fit = []
seq_fitness = []
seq_best_fit = []
best_fit = []
best_links = []
best_fit_keys = {}

for i, runNum in enumerate(c.runnumberstot):
    parents = pickle.load( open( f"pickles/{runNum}/save_parents{runNum}-500.p", "rb" ) )
    if runNum <= 5:
        for j,key in enumerate(parents.keys()):
            rand_fitness.append(parents[key].fitness)
            best_fit.append(parents[key].fitness)
            best_links.append(parents[key].numLinks)
            if j == 0:
                bestfit = parents[key].fitness
                bestfitkey = key
            if parents[key].fitness < bestfit:
                bestfit = parents[key].fitness
                bestfitkey = key
        rand_best_fit.append(bestfit)
    else:
        for j,key in enumerate(parents.keys()):
            seq_fitness.append(parents[key].fitness)
            best_fit.append(parents[key].fitness)
            best_links.append(parents[key].numLinks)
            if j == 0:
                bestfit = parents[key].fitness
                bestfitkey = key
            if parents[key].fitness < bestfit:
                bestfit = parents[key].fitness
                bestfitkey = key
        seq_best_fit.append(bestfit)
    best_fit_keys[runNum] = bestfitkey
    #best_fit.append(bestfit)
    #best_links.append(parents[bestfitkey].numLinks)
    
    
plt.figure()
plt.scatter(best_links,best_fit)
plt.xlabel("Number of Links")
plt.ylabel("Distance from Origin in -x Direction")
plt.savefig("Links-fitness-scatter.png")

corr_matrix = np.corrcoef(best_links, best_fit)
corr = corr_matrix[0,1]
R_sq = corr**2
 
print(R_sq)
            
            
print(ttest_ind(a=rand_fitness,b=seq_fitness,equal_var=False))
print(ttest_ind(a=rand_best_fit,b=seq_best_fit,equal_var=False))
#print(best_fit_keys)

fitness0 = {}

for i in range(1,11):
    fitness0[i] = []

for i, runNum in enumerate(c.runnumberstot):
    parents = pickle.load( open( f"pickles/{runNum}/save_parents{runNum}-0.p", "rb" ) )
    for j,key in enumerate(parents.keys()):
        fitness0[i+1].append(parents[key].fitness)

highlight = PatternFill(start_color='00FFFF00',end_color='00FFFF00',fill_type='solid')



wb = Workbook()

for i, runNum in enumerate(c.runnumberstot):
    mutations = pickle.load( open(f"pickles/{runNum}/save_mutations{runNum}-500-fit.p", "rb" ) )

    bestfitkey = best_fit_keys[runNum]
    df = pd.DataFrame(data=mutations)

    df0 = df[["generation",f"{bestfitkey}Success",f"{bestfitkey}Type",f"{bestfitkey}Specifics",f"{bestfitkey}Fit"]]

   
    
    df1 = df0[df0[f"{bestfitkey}Success"]=="True"].copy()
    
    df1['Diff'] = df1[f"{bestfitkey}Fit"].diff()*-1
    df1 = df1.reset_index(drop=True)

    df1.iloc[0,5] = -df1.iloc[0,4] + fitness0[runNum][bestfitkey]
    newsheet = f"{runNum}"
    ws = wb.create_sheet(newsheet)
    #df1 = df1.style.applymap(highlight_cells1)
    for r in dataframe_to_rows(df1, index=False, header=True):
        ws.append(r)
    
    for i in range(df1.shape[0]):
        if df1.iloc[i,5] > 0.1:
             ws[f"F{i+2}"].fill = highlight
    
    muts =[1,2,3,4,5]
    finalrow = df1.shape[0] + 1
    startrow = df1.shape[0] + 3
    for i, mut in enumerate(muts):
        ws[f"B{startrow+i}"] = mut
        ws[f"C{startrow+i}"] = f"=COUNTIF(C2:C{finalrow},{mut})"
        if i == len(muts)-1:
            ws[f"B{startrow+i+1}"] = "Total"
            ws[f"C{startrow+i+1}"] = f"=SUM(C{startrow}:C{startrow+len(muts)-1})"
        
    #df1.to_excel(f"pickles/{runNum}/mutationsTrue{runNum}.xlsx")
wb.save(f"pickles/mutationsTrue.xlsx")

